'''
Methods or writing the Excel output for a ReportWriter Report.
'''
import sys
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.styles import colors
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

MIN_WIDTH = 10
MAX_WIDTH = 30
MAX_CELL_WIDTH = 50
MAX_CODE_WIDTH = 10

def get_column_headings(line,delimiter):
    '''Return the column headings from the text output.

    :param str line: first line of the text file
    :param str delimiter: column delimiter (tab)

    :rtype: Array
    :returns: array containing the column headings
    '''
    return line.split(delimiter)

def get_max_column_lengths(input_file,delimiter):
    ''' Get the maximum length of the characters for each column

    :param File input_file: the text output file
    :param str delimiter: column delimiter (tab)

    :rtype: Array
    :returns: array containing the maximum lengths for each column
    '''

    column_lengths = []
    #
    # Read in the Header line and set initial max size
    #
    line = input_file.readline()
    line = line.strip()
    headings  = get_column_headings(line,delimiter)
    for i,heading in enumerate(headings):
       column_lengths.append(len(heading))

    for line in input_file:
        line = line.strip()
        columns = line.split(delimiter)
        for i,column in enumerate(columns):
            l = len(column)
            if l > column_lengths[i]:
                column_lengths[i] = l

    return column_lengths

def write_excel(input_file_name,report_columns):
    ''' Write Excel output based on the text output file

    :param File input_file: the text output file
    :param Array columns: array containing the maximum lengths for each column

    '''
    try:
        input_file = open(input_file_name,"r")
    except IOError:
        print("The input file does not exit")
        sys.exit(1)

    output_file = input_file_name.replace(".txt",".xlsx")
    #
    # Open Excel for writing
    #
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Report Writer"
    #
    # Read in the heading line
    #
    line = input_file.readline()
    line = line.strip()
    headings  = get_column_headings(line,"\t")
    input_file.seek(0)
    column_lengths = get_max_column_lengths(input_file,"\t")
    input_file.seek(0)

    col = 1
    row = 1
    headingFont = Font(color=colors.BLACK,bold=True)
    headingFill = PatternFill(fill_type='solid',fgColor="d3d3d3")
    headingAlignment = Alignment(vertical="center",horizontal="left")
    cellFont = Font(color=colors.BLACK,bold=False)
    cellAlignment = Alignment(vertical="top",horizontal="left",wrap_text=True)

    for heading in headings:
        sheet.cell(column=col, row=row).value = heading
        sheet.cell(column=col, row=row).font = headingFont
        sheet.cell(column=col, row=row).fill = headingFill
        sheet.cell(column=col, row=row).alignment = headingAlignment
        col = col + 1

    line = input_file.readline()
    for line in input_file:
        row = row + 1
        line = line.strip()
        columns = line.split("\t")
        for i,column in enumerate(columns):
            col = i + 1
            sheet.cell(column=col, row=row).value = column
            sheet.cell(column=col, row=row).font = cellFont
            sheet.cell(column=col, row=row).alignment = cellAlignment

    for i, column_length in enumerate(column_lengths):
        if column_length < MIN_WIDTH:
            sheet.column_dimensions[get_column_letter(i+1)].width = MIN_WIDTH
        elif column_length < MAX_CELL_WIDTH:
            sheet.column_dimensions[get_column_letter(i+1)].width = column_length
        else:
            sheet.column_dimensions[get_column_letter(i+1)].width = MAX_CELL_WIDTH

    wb.save(output_file)
    input_file.close()
